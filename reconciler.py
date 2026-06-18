"""
Inventory Reconciliation Engine v2
QBO vs TicketVault (Purchase Details + PO Cost Changes)
"""

import re
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Company mapping: QBO Company -> list of TV Company Names
# ---------------------------------------------------------------------------
COMPANY_MAPPING = {
    "Damona & Crew":      ["Damon and Crew"],
    "The Ticket Guy LLC": ["The Ticket Guy", "Ticket Guy"],
    "Y&S Tickets":        ["YS Tickets", "YS-SeatGeek2", "YS-Seatgeek", "YS Tickets Spec", "YS-Seatgeek2"],
    "YourTickets":        ["YourTickets"],
    "YS Asher Tickets":   ["YSA", "YSA 2", "YSA 3"],
    "YS Chase Tickets":   ["Jacks YS", "Chase (Jacks)"],
    "YS Katz Tickets":    ["YS Katz"],
    "YS Levine Tickets":  ["Yoni Levine"],
    "YS Levovitz Tickets":["Levovitz"],
    "YS Needle Tickets":  ["Needle Tickets LLC"],
    "YS TL Tickets":      ["YS TL"],
    "YSKG Tickets":       ["GK LLC", "YSKG"],
    "YSM Tickets":        ["YSM Tickets"],
    "YSP Tickets":        ["Pollak Tickets"],
    "YSS Tickets":        ["YSS Tickets"],
    "YSW Tickets":        ["YSW", "YSW (Waxler)"],
}

TV_TO_QBO = {}
for _qbo, _tv_list in COMPANY_MAPPING.items():
    for _tv in _tv_list:
        TV_TO_QBO[_tv.strip().lower()] = _qbo

DESC_TO_QBO = {}
for _qbo, _tv_list in COMPANY_MAPPING.items():
    for _tv in _tv_list:
        DESC_TO_QBO[_tv.strip().lower()] = _qbo
    DESC_TO_QBO[_qbo.strip().lower()] = _qbo

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _extract_desc_company(description):
    if not description or pd.isna(description):
        return None
    m = re.search(r'\(([^)]+)\)\s*$', str(description).strip())
    return m.group(1).strip() if m else None

def _map_tv_to_qbo(tv_company):
    return TV_TO_QBO.get(str(tv_company).strip().lower())

def _map_desc_to_qbo(desc_company):
    return DESC_TO_QBO.get(str(desc_company).strip().lower())

def _clean_num(val):
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return 'N/A' if s.lower() in ('nan', 'none', '') else s

def _fmt_date(val):
    try:
        ts = pd.Timestamp(val)
        if pd.isna(ts):
            return ''
        return ts.strftime('%m/%d/%Y')
    except:
        return ''

# ---------------------------------------------------------------------------
# QBO Parsing
# ---------------------------------------------------------------------------
def parse_qbo_consolidated(filepath_or_buffer):
    raw = pd.read_excel(filepath_or_buffer, header=None)
    header_row = None
    for i, row in raw.iterrows():
        if any(str(v).strip().lower() == 'transaction date' for v in row if pd.notna(v)):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find header row in consolidated QBO file.")

    df = pd.read_excel(filepath_or_buffer, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if 'company' in cl and 'transaction' not in cl: col_map['qbo_company'] = c
        elif 'transaction date' in cl: col_map['date'] = c
        elif 'transaction type' in cl: col_map['transaction_type'] = c
        elif cl == 'num': col_map['num'] = c
        elif cl == 'name': col_map['name'] = c
        elif 'description' in cl and 'full' not in cl: col_map['description'] = c
        elif 'amount' in cl and 'balance' not in cl: col_map['amount'] = c

    missing = [r for r in ['qbo_company','date','transaction_type','num','description','amount'] if r not in col_map]
    if missing:
        raise ValueError(f"Consolidated QBO missing columns: {missing}")

    df = df.rename(columns={v: k for k, v in col_map.items()})
    final_cols = ['qbo_company','transaction_type','date','num','name','description','amount']
    for c in final_cols:
        if c not in df.columns: df[c] = None
    df = df[final_cols].copy()

    df = df[df['qbo_company'].notna() & (df['qbo_company'].astype(str).str.strip() != '')]
    df = df[df['date'].notna()]
    df = df[~df['qbo_company'].astype(str).str.strip().str.lower().isin(['company','beginning balance',''])]
    df = df[df['transaction_type'].notna()]
    df = df[~df['transaction_type'].astype(str).str.strip().str.lower().isin(['nan','none',''])]

    df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.normalize()
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
    df['num'] = df['num'].astype(str).str.strip()
    df['qbo_company'] = df['qbo_company'].astype(str).str.strip()
    df['transaction_type'] = df['transaction_type'].astype(str).str.strip()
    df['source_file'] = 'consolidated'
    return df.dropna(subset=['date'])


def parse_qbo_single(filepath_or_buffer, filename=''):
    raw = pd.read_excel(filepath_or_buffer, header=None)
    company_name = None
    for i in range(5):
        val = str(raw.iloc[i, 0]).strip()
        if val and val.lower() not in ['nan','none','','daily inventory summary']:
            company_name = val
            break
    if not company_name:
        import os
        company_name = os.path.splitext(os.path.basename(filename))[0]

    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if pd.notna(v)]
        if 'transaction date' in vals or ('num' in vals and 'amount' in vals):
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"Could not find header row for '{company_name}'.")

    df = pd.read_excel(filepath_or_buffer, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    first_col = df.columns[0]
    has_txn_type_col = any('transaction type' in c.lower() for c in df.columns)

    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if 'transaction date' in cl: col_map['date'] = c
        elif 'transaction type' in cl: col_map['transaction_type'] = c
        elif cl == 'num': col_map['num'] = c
        elif cl == 'name': col_map['name'] = c
        elif 'description' in cl and 'full' not in cl: col_map['description'] = c
        elif 'amount' in cl and 'balance' not in cl: col_map['amount'] = c

    if 'date' not in col_map:
        raise ValueError(f"Could not find 'Transaction date' column for '{company_name}'.")

    if not has_txn_type_col:
        txn_types = []
        current_type = 'Unknown'
        known_types = {'bill','expense','check','journal entry','credit card credit'}
        for val in df[first_col]:
            v = str(val).strip()
            if v.lower() in known_types:
                current_type = v.title()
                txn_types.append(None)
            else:
                txn_types.append(current_type)
        df['transaction_type'] = txn_types
        col_map['transaction_type'] = 'transaction_type'

    df['qbo_company'] = company_name
    df = df.rename(columns={v: k for k, v in col_map.items()})
    if first_col in df.columns and first_col not in ('transaction_type','date','num','name','description','amount','qbo_company'):
        df = df.drop(columns=[first_col])

    final_cols = ['qbo_company','transaction_type','date','num','name','description','amount']
    for c in final_cols:
        if c not in df.columns: df[c] = None
    df = df[final_cols].copy()

    df = df[df['date'].notna()]
    df = df[~df['date'].astype(str).str.strip().str.lower().isin(['nan','none','transaction date','beginning balance',''])]
    df = df[df['transaction_type'].notna()]
    df = df[~df['transaction_type'].astype(str).str.strip().str.lower().isin(['nan','none','','unknown'])]

    df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.normalize()
    df = df.dropna(subset=['date'])
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
    df['num'] = df['num'].astype(str).str.strip()
    df['qbo_company'] = df['qbo_company'].astype(str).str.strip()
    df['transaction_type'] = df['transaction_type'].astype(str).str.strip()
    df['source_file'] = company_name
    return df


# ---------------------------------------------------------------------------
# TV Parsing
# ---------------------------------------------------------------------------
def parse_purchase_details(filepath_or_buffers):
    """
    Parse Purchase Details file(s) — uses the 'All' tab.
    Returns (recon_df, raw_df).
    recon_df: company, date, total_cost grouped for reconciliation.
    raw_df: full cleaned dataframe for output tab.
    """
    if not isinstance(filepath_or_buffers, list):
        filepath_or_buffers = [filepath_or_buffers]

    frames = []
    for buf in filepath_or_buffers:
        try:
            df = pd.read_excel(buf, sheet_name='All')
        except:
            df = pd.read_excel(buf)
        df.columns = [str(c).strip() for c in df.columns]
        frames.append(df)

    df = pd.concat(frames, ignore_index=True)

    # Normalize date
    df['PO Created'] = pd.to_datetime(df['PO Created'], errors='coerce').dt.normalize()
    df['tv_company'] = df['Company'].astype(str).str.strip()
    df['total_cost'] = pd.to_numeric(df['Total Cost'], errors='coerce').fillna(0)
    df['qbo_company'] = df['tv_company'].apply(_map_tv_to_qbo)

    raw_df = df.copy()

    recon_df = df[['tv_company','qbo_company','PO Created','total_cost']].copy()
    recon_df = recon_df.rename(columns={'PO Created': 'date'})
    recon_df = recon_df.dropna(subset=['date'])

    return recon_df, raw_df


def parse_po_cost_changes(filepath_or_buffers):
    """
    Parse PO Cost Changes file(s) — uses the 'Combined' tab.
    Positive amounts = Bills, negative = Expenses.
    Returns (recon_df, raw_df).
    """
    if not isinstance(filepath_or_buffers, list):
        filepath_or_buffers = [filepath_or_buffers]

    frames = []
    for buf in filepath_or_buffers:
        try:
            df = pd.read_excel(buf, sheet_name='Combined')
        except:
            df = pd.read_excel(buf)
        df.columns = [str(c).strip() for c in df.columns]
        frames.append(df)

    df = pd.concat(frames, ignore_index=True)

    # Date column is 'Date'
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.normalize()
    df['tv_company'] = df['Company'].astype(str).str.strip()

    # Amount column is 'Total'
    df['total'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df['qbo_company'] = df['tv_company'].apply(_map_tv_to_qbo)

    raw_df = df.copy()

    recon_df = df[['tv_company','qbo_company','Date','total']].copy()
    recon_df = recon_df.rename(columns={'Date': 'date', 'total': 'amount'})
    recon_df = recon_df.dropna(subset=['date'])

    return recon_df, raw_df


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------

def _build_recon(qbo_left, tv_right, qbo_amount_col='amount', tv_amount_col='amount'):
    """Generic daily recon: group both sides by company+date and merge."""
    qbo_g = qbo_left.groupby(['qbo_company','date'])[qbo_amount_col].sum().reset_index().rename(columns={qbo_amount_col:'qbo_total'})
    tv_g  = tv_right.groupby(['qbo_company','date'])[tv_amount_col].sum().reset_index().rename(columns={tv_amount_col:'tv_total'})
    merged = pd.merge(qbo_g, tv_g, on=['qbo_company','date'], how='outer').fillna(0)
    merged['variance'] = merged['qbo_total'] - merged['tv_total']
    merged['match'] = merged['variance'].abs() <= 100.00
    return merged.sort_values(['qbo_company','date'])


def check1a_bills_recon(qbo_df, pd_recon_df, cc_recon_df):
    """
    1a — Bills: QBO Bills vs (Purchase Details + PO Cost Changes positives).
    """
    qbo_bills = qbo_df[qbo_df['transaction_type'].str.lower() == 'bill'].copy()

    # TV Bills = Purchase Details + positive PO Cost Changes
    tv_pd = pd_recon_df[['qbo_company','date','total_cost']].copy().rename(columns={'total_cost':'amount'})
    tv_cc_pos = cc_recon_df[cc_recon_df['amount'] > 0][['qbo_company','date','amount']].copy()
    tv_bills = pd.concat([tv_pd, tv_cc_pos], ignore_index=True)
    tv_bills = tv_bills[tv_bills['qbo_company'].notna()]

    result = _build_recon(qbo_bills, tv_bills)

    # Unmapped TV companies
    unmapped_pd = pd_recon_df[pd_recon_df['qbo_company'].isna()][['tv_company','date','total_cost']].rename(columns={'total_cost':'amount'})
    unmapped_cc = cc_recon_df[(cc_recon_df['qbo_company'].isna()) & (cc_recon_df['amount'] > 0)][['tv_company','date','amount']]
    unmapped = pd.concat([unmapped_pd, unmapped_cc], ignore_index=True)
    unmapped = unmapped.groupby(['tv_company','date'])['amount'].sum().reset_index()

    return result, unmapped


def check1b_expenses_recon(qbo_df, cc_recon_df):
    """
    1b — Expenses: QBO Expenses vs PO Cost Changes negatives.
    """
    qbo_expenses = qbo_df[qbo_df['transaction_type'].str.lower() == 'expense'].copy()

    # TV Expenses = negative PO Cost Changes (negate to make positive for comparison)
    tv_exp = cc_recon_df[cc_recon_df['amount'] < 0][['qbo_company','date','amount']].copy()
    tv_exp['amount'] = tv_exp['amount'].abs()
    tv_exp = tv_exp[tv_exp['qbo_company'].notna()]

    # QBO expense amounts are typically negative — negate for comparison
    qbo_expenses = qbo_expenses.copy()
    qbo_expenses['amount_abs'] = qbo_expenses['amount'].abs()

    qbo_g = qbo_expenses.groupby(['qbo_company','date'])['amount_abs'].sum().reset_index().rename(columns={'amount_abs':'qbo_total'})
    tv_g  = tv_exp.groupby(['qbo_company','date'])['amount'].sum().reset_index().rename(columns={'amount':'tv_total'})
    merged = pd.merge(qbo_g, tv_g, on=['qbo_company','date'], how='outer').fillna(0)
    merged['variance'] = merged['qbo_total'] - merged['tv_total']
    merged['match'] = merged['variance'].abs() <= 100.00
    return merged.sort_values(['qbo_company','date'])


def check1c_combined_recon(qbo_df, pd_recon_df, cc_recon_df):
    """
    1c — Combined: QBO (Bills + Expenses) vs all TV activity.
    QBO expenses are negative; TV expenses (negative CC) are also negative — sum all.
    """
    qbo_be = qbo_df[qbo_df['transaction_type'].str.lower().isin(['bill','expense'])].copy()

    # TV combined = Purchase Details (positive) + PO Cost Changes (positive and negative)
    tv_pd = pd_recon_df[['qbo_company','date','total_cost']].copy().rename(columns={'total_cost':'amount'})
    tv_cc = cc_recon_df[['qbo_company','date','amount']].copy()
    tv_all = pd.concat([tv_pd, tv_cc], ignore_index=True)
    tv_all = tv_all[tv_all['qbo_company'].notna()]

    result = _build_recon(qbo_be, tv_all)

    unmapped_pd = pd_recon_df[pd_recon_df['qbo_company'].isna()][['tv_company','date','total_cost']].rename(columns={'total_cost':'amount'})
    unmapped_cc = cc_recon_df[cc_recon_df['qbo_company'].isna()][['tv_company','date','amount']]
    unmapped = pd.concat([unmapped_pd, unmapped_cc], ignore_index=True)
    unmapped = unmapped.groupby(['tv_company','date'])['amount'].sum().reset_index()

    return result, unmapped


def check2_duplicate_bills(qbo_df):
    """2a: Same Company + Bill # (any transaction type)."""
    txns = qbo_df[qbo_df['transaction_type'].str.lower().isin(['bill','expense'])].copy()
    txns = txns[txns['num'].notna() & (txns['num'] != '') & (txns['num'].str.lower() != 'nan')]
    dupes = txns[txns.duplicated(subset=['qbo_company','num'], keep=False)].copy()
    return dupes.sort_values(['qbo_company','num','date'])


def check2b_duplicate_bills_detail(qbo_df):
    """2b: Same Company + Date + Name + Description + Amount."""
    txns = qbo_df[qbo_df['transaction_type'].str.lower().isin(['bill','expense'])].copy()
    txns['_name_n'] = txns['name'].astype(str).str.strip().str.lower()
    txns['_desc_n'] = txns['description'].astype(str).str.strip().str.lower()
    dupes = txns[txns.duplicated(subset=['qbo_company','date','_name_n','_desc_n','amount'], keep=False)].copy()
    dupes = dupes.drop(columns=['_name_n','_desc_n'])
    return dupes.sort_values(['qbo_company','date','name'])


def check3_description_mismatch(qbo_df):
    relevant = qbo_df[qbo_df['transaction_type'].str.lower().isin(['bill','expense'])].copy()
    relevant['desc_company_raw'] = relevant['description'].apply(_extract_desc_company)
    relevant['desc_qbo_mapped'] = relevant['desc_company_raw'].apply(lambda x: _map_desc_to_qbo(x) if x else None)

    # Only flag rows where the text in parentheses IS a recognized company name
    # (i.e. desc_qbo_mapped is not None) but maps to a DIFFERENT QBO company
    # Rows where the parentheses text is not a known company name are ignored
    mismatches = relevant[
        relevant['desc_company_raw'].notna() &
        relevant['desc_qbo_mapped'].notna() &
        (relevant['desc_qbo_mapped'].str.lower() != relevant['qbo_company'].str.lower())
    ].copy()
    mismatches['mismatch_reason'] = 'Expected: ' + mismatches['desc_qbo_mapped'] + ' | Got: ' + mismatches['qbo_company']

    return mismatches.sort_values(['qbo_company','date'])


# ---------------------------------------------------------------------------
# Excel Styles
# ---------------------------------------------------------------------------
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")
GREEN_FILL  = PatternFill("solid", fgColor="CCFFCC")
ORANGE_FILL = PatternFill("solid", fgColor="FFE0B2")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL    = PatternFill("solid", fgColor="2F5496")
WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BOLD_FONT   = Font(bold=True, name="Calibri", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)


def _style_header(ws, row, cols, fill=HEADER_FILL):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        header_len = len(str(col[0].value)) if col[0].value else 0
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, header_len + 3, max_len + 2))


def _write_recon_sheet(ws, recon_df, unmapped_df, qbo_label, tv_label):
    """Write a standard reconciliation sheet (1a/1b/1c)."""
    headers = ["QBO Company", "Date", f"QBO {qbo_label} ($)", f"TV {tv_label} ($)", "Variance ($)", "Status"]
    for c, h in enumerate(headers, 1): ws.cell(1, c, h)
    _style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(recon_df.itertuples(), 2):
        ws.cell(r, 1, row.qbo_company).font = NORMAL_FONT
        ws.cell(r, 2, _fmt_date(row.date)).font = NORMAL_FONT
        ws.cell(r, 3, round(row.qbo_total, 2)).number_format = '$#,##0.00'
        ws.cell(r, 4, round(row.tv_total, 2)).number_format = '$#,##0.00'
        ws.cell(r, 5, round(row.variance, 2)).number_format = '$#,##0.00'
        qbo_zero = abs(row.qbo_total) < 0.02
        tv_zero  = abs(row.tv_total)  < 0.02
        if row.match:
            label, fill = "Match", GREEN_FILL
        elif qbo_zero or tv_zero:
            label, fill = "Missing", RED_FILL
        else:
            label, fill = "Discrepancy", ORANGE_FILL
        ws.cell(r, 6, label).font = NORMAL_FONT
        for c in range(1, 7): ws.cell(r, c).fill = fill

    if unmapped_df is not None and len(unmapped_df) > 0:
        start = len(recon_df) + 3
        ws.cell(start, 1, "⚠️ Unmapped TV Companies").font = Font(bold=True, color="CC6600", name="Calibri", size=10)
        ws.merge_cells(f"A{start}:F{start}")
        for c, h in enumerate(["TV Company","Date","TV Total ($)"], 1): ws.cell(start+1, c, h)
        _style_header(ws, start+1, 3, fill=SUB_FILL)
        for r2, row2 in enumerate(unmapped_df.itertuples(), start+2):
            ws.cell(r2, 1, row2.tv_company).font = NORMAL_FONT
            ws.cell(r2, 2, _fmt_date(row2.date)).font = NORMAL_FONT
            ws.cell(r2, 3, round(row2.amount, 2)).number_format = '$#,##0.00'
            for c in range(1, 4): ws.cell(r2, c).fill = ORANGE_FILL

    _auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(recon_df)+1}"


def _write_dupes_sheet(ws, df, headers, col_fns):
    if len(df) == 0:
        ws["A1"] = "No duplicates found."
        ws["A1"].font = Font(bold=True, color="006600", name="Calibri", size=12)
        return
    for c, h in enumerate(headers, 1): ws.cell(1, c, h)
    _style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22
    amt_col = len(col_fns)
    for r, row in enumerate(df.itertuples(), 2):
        for c, fn in enumerate(col_fns, 1):
            cell = ws.cell(r, c, fn(row))
            cell.font = NORMAL_FONT
            if c == amt_col:
                cell.number_format = '$#,##0.00'
    _auto_width(ws, max_w=60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_raw_tv_sheet(ws, df, date_cols, money_cols, drop_cols=None):
    """Write a raw TV dataframe to a sheet with date/money formatting."""
    if drop_cols:
        df = df.drop(columns=[c for c in drop_cols if c in df.columns])
    # Drop internal helper columns
    cols = [c for c in df.columns if not c.startswith('_') and c not in ('tv_company','total_cost','qbo_company','total','amount')]
    for c, h in enumerate(cols, 1): ws.cell(1, c, h)
    _style_header(ws, 1, len(cols))
    ws.row_dimensions[1].height = 22
    for r, row in enumerate(df[cols].itertuples(index=False), 2):
        for c, (col_name, val) in enumerate(zip(cols, row), 1):
            if col_name in date_cols:
                ws.cell(r, c, _fmt_date(val)).font = NORMAL_FONT
            elif col_name in money_cols:
                cell = ws.cell(r, c, val)
                cell.number_format = '$#,##0.00'
                cell.font = NORMAL_FONT
            else:
                ws.cell(r, c, val if val is not None and str(val) != 'nan' else '').font = NORMAL_FONT
    # Header-width columns for TV tabs
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header_len = len(str(col[0].value)) if col[0].value else 10
        ws.column_dimensions[col_letter].width = header_len + 3
    ws.freeze_panes = "A2"
    if len(df) > 0: ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Report Builder
# ---------------------------------------------------------------------------
def build_report(qbo_df, pd_recon_df, pd_raw_df, cc_recon_df, cc_raw_df,
                 period_label="", input_files=None):
    wb = Workbook()

    # Run all checks
    recon_1a, unmapped_1a = check1a_bills_recon(qbo_df, pd_recon_df, cc_recon_df)
    recon_1b             = check1b_expenses_recon(qbo_df, cc_recon_df)
    recon_1c, unmapped_1c = check1c_combined_recon(qbo_df, pd_recon_df, cc_recon_df)
    dupes_df             = check2_duplicate_bills(qbo_df)
    dupes2_df            = check2b_duplicate_bills_detail(qbo_df)
    mismatch_df          = check3_description_mismatch(qbo_df)

    # ── Summary ─────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    SUMMARY_FILL = PatternFill("solid", fgColor="EEF2FA")
    SUMMARY_ITEMS = [
        ("Tab", "What it checks", "Flag criteria"),
        ("1a. Bills Recon", "QBO Bills vs TicketVault Purchase Details + positive PO Cost Changes, grouped by date & company.", "Match (green) = within $100  |  Missing (red) = one side is $0  |  Discrepancy (orange) = both have $ but differ by more than $100"),
        ("1b. Expenses Recon", "QBO Expenses vs negative PO Cost Changes (absolute values compared), grouped by date & company.", "Same color logic as above"),
        ("1c. Combined Recon", "QBO Bills + Expenses vs all TV activity (Purchase Details + PO Cost Changes), grouped by date & company.", "Same color logic as above"),
        ("2. Duplicate Bills (Bill #s)", "Bills or Expenses where the same Company + Bill # appears more than once in QBO.", "Any Company + Bill # combination appearing 2 or more times"),
        ("3. Duplicate Bills (Detail)", "Bills or Expenses where Company, Date, Name, Description, and Amount all match.", "All five fields must match across two or more rows"),
        ("4. Description Mismatches", "For Bills and Expenses, checks that the company in parentheses in the Description matches the QBO Company column. Only flags when the parentheses text is a recognized company name.", "Company in parentheses maps to a different QBO company than the Company column"),
    ]
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 80
    ws_sum.column_dimensions['C'].width = 55
    for r, (tab, desc, criteria) in enumerate(SUMMARY_ITEMS, 1):
        for c, val in enumerate([tab, desc, criteria], 1):
            cell = ws_sum.cell(r, c, val)
            cell.alignment = Alignment(wrap_text=False, vertical='center')
            if r == 1:
                cell.fill = HEADER_FILL
                cell.font = WHITE_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.fill = SUMMARY_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                cell.font = BOLD_FONT if c == 1 else NORMAL_FONT
        ws_sum.row_dimensions[r].height = 20

    # ── 1a. Bills Recon ─────────────────────────────────────────────────────
    ws1a = wb.create_sheet("1a. Bills Recon")
    _write_recon_sheet(ws1a, recon_1a, unmapped_1a, "Bills Total", "Bills Total")

    # ── 1b. Expenses Recon ──────────────────────────────────────────────────
    ws1b = wb.create_sheet("1b. Expenses Recon")
    _write_recon_sheet(ws1b, recon_1b, None, "Expenses Total", "Cost Changes Total")

    # ── 1c. Combined Recon ──────────────────────────────────────────────────
    ws1c = wb.create_sheet("1c. Combined Recon")
    _write_recon_sheet(ws1c, recon_1c, unmapped_1c, "Bills + Expenses", "All TV Activity")

    # ── 2. Duplicate Bills (Bill #s) ────────────────────────────────────────
    ws2 = wb.create_sheet("2. Duplicate Bills (Bill #s)")
    _write_dupes_sheet(ws2, dupes_df,
        ["QBO Company","Transaction Type","Bill / Expense #","Date","Name","Description","Amount ($)"],
        [lambda r: r.qbo_company, lambda r: r.transaction_type, lambda r: _clean_num(r.num),
         lambda r: _fmt_date(r.date), lambda r: str(r.name) if pd.notna(r.name) else '',
         lambda r: str(r.description) if pd.notna(r.description) else '', lambda r: round(r.amount, 2)])

    # ── 3. Duplicate Bills (Detail) ─────────────────────────────────────────
    ws3 = wb.create_sheet("3. Duplicate Bills (Detail)")
    _write_dupes_sheet(ws3, dupes2_df,
        ["QBO Company","Transaction Type","Date","Bill / Expense #","Name","Description","Amount ($)"],
        [lambda r: r.qbo_company, lambda r: r.transaction_type, lambda r: _fmt_date(r.date),
         lambda r: _clean_num(r.num), lambda r: str(r.name) if pd.notna(r.name) else '',
         lambda r: str(r.description) if pd.notna(r.description) else '', lambda r: round(r.amount, 2)])

    # ── 4. Description Mismatches ───────────────────────────────────────────
    ws4 = wb.create_sheet("4. Description Mismatches")
    if len(mismatch_df) == 0:
        ws4["A1"] = "No description company mismatches found."
        ws4["A1"].font = Font(bold=True, color="006600", name="Calibri", size=12)
    else:
        h4 = ["QBO Company","Transaction Type","Date","Bill / Expense #","Description","Mismatch Detail","Amount ($)"]
        for c, h in enumerate(h4, 1): ws4.cell(1, c, h)
        _style_header(ws4, 1, len(h4))
        ws4.row_dimensions[1].height = 22
        for r, row in enumerate(mismatch_df.itertuples(), 2):
            ws4.cell(r, 1, row.qbo_company).font = NORMAL_FONT
            ws4.cell(r, 2, row.transaction_type).font = NORMAL_FONT
            ws4.cell(r, 3, _fmt_date(row.date)).font = NORMAL_FONT
            ws4.cell(r, 4, _clean_num(row.num)).font = NORMAL_FONT
            ws4.cell(r, 5, str(row.description) if pd.notna(row.description) else '').font = NORMAL_FONT
            ws4.cell(r, 6, str(row.mismatch_reason)).font = NORMAL_FONT
            ws4.cell(r, 7, round(row.amount, 2)).number_format = '$#,##0.00'
    _auto_width(ws4, max_w=60)
    ws4.freeze_panes = "A2"
    if len(mismatch_df) > 0: ws4.auto_filter.ref = ws4.dimensions

    # ── TV - Purchase Details (raw) ─────────────────────────────────────────
    ws_pd = wb.create_sheet("TV - Purchase Details")
    _write_raw_tv_sheet(ws_pd, pd_raw_df,
        date_cols={'PO Created'},
        money_cols={'Total Cost'})

    # ── TV - PO Cost Changes (raw) ──────────────────────────────────────────
    ws_cc = wb.create_sheet("TV - PO Cost Changes")
    _write_raw_tv_sheet(ws_cc, cc_raw_df,
        date_cols={'Date'},
        money_cols={'Total'})

    # ── QBO - Bills ─────────────────────────────────────────────────────────
    QBO_HEADERS = ["Company","Transaction Type","Transaction Date","Num","Name","Description","Amount ($)"]
    def _write_qbo_tab(ws, df_subset):
        for c, h in enumerate(QBO_HEADERS, 1): ws.cell(1, c, h)
        _style_header(ws, 1, len(QBO_HEADERS))
        ws.row_dimensions[1].height = 22
        for r, row in enumerate(df_subset.itertuples(), 2):
            ws.cell(r, 1, row.qbo_company).font = NORMAL_FONT
            ws.cell(r, 2, row.transaction_type).font = NORMAL_FONT
            ws.cell(r, 3, _fmt_date(row.date)).font = NORMAL_FONT
            ws.cell(r, 4, _clean_num(row.num)).font = NORMAL_FONT
            ws.cell(r, 5, str(row.name) if pd.notna(row.name) else '').font = NORMAL_FONT
            ws.cell(r, 6, str(row.description) if pd.notna(row.description) else '').font = NORMAL_FONT
            cell = ws.cell(r, 7, round(row.amount, 2))
            cell.number_format = '$#,##0.00'
            cell.font = NORMAL_FONT
        _auto_width(ws, max_w=60)
        ws.freeze_panes = "A2"
        if len(df_subset) > 0: ws.auto_filter.ref = ws.dimensions

    qbo_excl_je = qbo_df[~qbo_df['transaction_type'].str.lower().isin(['journal entry'])]
    ws_qbo_b = wb.create_sheet("QBO - Bills")
    _write_qbo_tab(ws_qbo_b, qbo_excl_je[qbo_excl_je['transaction_type'].str.lower() == 'bill'])
    ws_qbo_e = wb.create_sheet("QBO - Expenses")
    _write_qbo_tab(ws_qbo_e, qbo_excl_je[qbo_excl_je['transaction_type'].str.lower() != 'bill'])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
